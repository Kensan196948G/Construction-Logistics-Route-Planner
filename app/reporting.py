from __future__ import annotations

import csv
import logging
from io import BytesIO, StringIO

from app.models import (
    DISCLAIMER,
    SAMPLE_DATA_NOTICE,
    Project,
    RiskLevel,
    RouteCandidate,
)
from app.risk_engine import risk_counts

logger = logging.getLogger(__name__)


LEVEL_LABELS = {
    RiskLevel.candidate: "利用候補",
    RiskLevel.caution: "注意",
    RiskLevel.confirm_required: "要確認",
    RiskLevel.exclusion_consideration: "除外検討",
    RiskLevel.data_insufficient: "データ不足",
}


def render_markdown(project: Project, routes: list[RouteCandidate]) -> str:
    lines = [
        "# 搬入ルート初期検討メモ",
        "",
        "> ⚠️ **本番利用禁止（PoC・サンプル）**",
        f"> {SAMPLE_DATA_NOTICE}",
        "",
        "## 1. 案件概要",
        "",
        f"- 工事件名: {project.project_name}",
        f"- 現場名: {project.site_name}",
        f"- 担当者: {project.planner}",
        f"- 発注者区分: {project.owner_type or '未入力'}",
        f"- 出発地: {project.start.name} ({project.start.lat:.6f}, {project.start.lng:.6f})",
        f"- 到着地: {project.destination.name} ({project.destination.lat:.6f}, {project.destination.lng:.6f})",
        "",
        "## 2. 搬入条件",
        "",
        f"- 車両種別: {project.vehicle.vehicle_type}",
        f"- 全長/全幅/全高: {_value(project.vehicle.length_m, 'm')} / {_value(project.vehicle.width_m, 'm')} / {_value(project.vehicle.height_m, 'm')}",
        f"- 総重量/軸重: {_value(project.vehicle.gross_weight_t, 't')} / {_value(project.vehicle.axle_weight_t, 't')}",
        f"- 積載物: {project.vehicle.cargo_type or '未入力'}",
        f"- 特車該当可能性: {'あり' if project.vehicle.special_vehicle_flag else '未指定'}",
        f"- 搬入日: {project.delivery.delivery_date or '未指定'}",
        f"- 時間帯: {project.delivery.time_window}",
        "",
        "## 3. ルート候補比較",
        "",
        "| 候補 | 距離 | 時間 | 評価 | スコア | 注意 | 要確認 | データ不足 | コメント |",
        "|---|---:|---:|---|---:|---:|---:|---:|---|",
    ]

    for route in routes:
        counts = risk_counts(route.risks)
        lines.append(
            "| "
            f"{route.name} | "
            f"{route.distance_km:.1f} km | "
            f"{route.duration_min} 分 | "
            f"{LEVEL_LABELS[route.risk_level]} | "
            f"{route.risk_score} | "
            f"{counts[RiskLevel.caution.value]} | "
            f"{counts[RiskLevel.confirm_required.value]} | "
            f"{counts[RiskLevel.data_insufficient.value]} | "
            f"{route.summary} |"
        )

    lines.extend(["", "## 4. 主な注意箇所", ""])
    for route in routes:
        lines.extend([f"### {route.name}", ""])
        if not route.risks:
            lines.append("- 評価済みの注意箇所はありません。ただし正式確認は必要です。")
            continue
        for risk in route.risks:
            location = ""
            if risk.feature:
                location = f" / 位置: {risk.feature.lat:.6f}, {risk.feature.lng:.6f}"
            lines.append(
                f"- [{LEVEL_LABELS[risk.level]}] {risk.title}: {risk.message} "
                f"確認先: {risk.confirmation_target}. 根拠: {risk.evidence}{location}"
            )

    lines.extend(
        [
            "",
            "## 5. 追加確認事項",
            "",
            "- 道路管理者へ橋梁、幅員、高さ、重量、時間帯規制の確認",
            "- 警察または関係機関へ道路使用、誘導、時間帯条件の確認",
            "- 協力会社へ車両諸元、待機場所、転回可否、過去通行実績の確認",
            "- 現地踏査で学校、病院、住宅地、交差点、踏切、急勾配の確認",
            "",
            "## 6. 注意文",
            "",
            DISCLAIMER,
        ]
    )
    return "\n".join(lines) + "\n"


def render_csv(project: Project, routes: list[RouteCandidate]) -> str:
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "project_id",
            "project_name",
            "route_id",
            "route_name",
            "distance_km",
            "duration_min",
            "risk_level",
            "risk_score",
            "risk_title",
            "risk_message",
            "confirmation_target",
            "evidence",
            "sample_notice",
        ]
    )
    for route in routes:
        if not route.risks:
            writer.writerow(
                [
                    project.id,
                    _csv_safe(project.project_name),
                    route.id,
                    _csv_safe(route.name),
                    route.distance_km,
                    route.duration_min,
                    route.risk_level.value,
                    route.risk_score,
                    "",
                    "",
                    "",
                    "",
                    SAMPLE_DATA_NOTICE,
                ]
            )
            continue
        for risk in route.risks:
            writer.writerow(
                [
                    project.id,
                    _csv_safe(project.project_name),
                    route.id,
                    _csv_safe(route.name),
                    route.distance_km,
                    route.duration_min,
                    route.risk_level.value,
                    route.risk_score,
                    _csv_safe(risk.title),
                    _csv_safe(risk.message),
                    _csv_safe(risk.confirmation_target),
                    _csv_safe(risk.evidence),
                    SAMPLE_DATA_NOTICE,
                ]
            )
    return output.getvalue()


def _csv_safe(value: str) -> str:
    """Neutralize spreadsheet formula injection in exported CSV cells.

    Cells beginning with =, +, -, @, tab or CR can be interpreted as formulas
    when the CSV is opened in Excel/LibreOffice. Prefixing a single quote
    keeps the value readable while preventing formula execution.
    """

    text = str(value)
    if text and text[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + text
    return text


def render_pdf(project: Project, routes: list[RouteCandidate]) -> bytes:
    """Render the initial-review memo as a PDF (reportlab, CID Japanese font)."""

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    try:
        pdfmetrics.registerFont(UnicodeCIDFont("HeiseiMin-W3"))
    except Exception:
        logger.warning("HeiseiMin-W3 CID font unavailable; PDF uses reportlab default.", exc_info=True)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitleJP", parent=styles["Title"], fontName="HeiseiMin-W3", fontSize=16, leading=22
    )
    body_style = ParagraphStyle(
        "BodyJP", parent=styles["BodyText"], fontName="HeiseiMin-W3", fontSize=9.5, leading=14
    )
    heading_style = ParagraphStyle(
        "HeadingJP", parent=styles["Heading2"], fontName="HeiseiMin-W3", fontSize=12, leading=16
    )
    warning_style = ParagraphStyle(
        "WarningJP",
        parent=body_style,
        textColor=colors.HexColor("#b3261e"),
        fontSize=9,
        leading=13,
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=18 * mm, rightMargin=18 * mm)
    story = [
        Paragraph("搬入ルート初期検討メモ", title_style),
        Spacer(1, 6),
        Paragraph(f"<b>本番利用禁止（PoC・サンプル）</b><br/>{SAMPLE_DATA_NOTICE}", warning_style),
        Spacer(1, 10),
        Paragraph("1. 案件概要", heading_style),
        Paragraph(
            f"工事件名: {project.project_name}<br/>"
            f"現場名: {project.site_name}<br/>"
            f"担当者: {project.planner}<br/>"
            f"発注者区分: {project.owner_type or '未入力'}<br/>"
            f"出発地: {project.start.name} ({project.start.lat:.6f}, {project.start.lng:.6f})<br/>"
            f"到着地: {project.destination.name} ({project.destination.lat:.6f}, {project.destination.lng:.6f})",
            body_style,
        ),
        Spacer(1, 8),
        Paragraph("2. 搬入条件", heading_style),
        Paragraph(
            f"車両種別: {project.vehicle.vehicle_type}<br/>"
            f"全長/全幅/全高: {_value(project.vehicle.length_m, 'm')} / "
            f"{_value(project.vehicle.width_m, 'm')} / {_value(project.vehicle.height_m, 'm')}<br/>"
            f"総重量/軸重: {_value(project.vehicle.gross_weight_t, 't')} / "
            f"{_value(project.vehicle.axle_weight_t, 't')}<br/>"
            f"積載物: {project.vehicle.cargo_type or '未入力'}<br/>"
            f"搬入日: {project.delivery.delivery_date or '未指定'} / 時間帯: {project.delivery.time_window}",
            body_style,
        ),
        Spacer(1, 8),
        Paragraph("3. ルート候補比較", heading_style),
    ]

    table_rows = [["候補", "距離", "時間", "評価", "スコア", "要確認", "データ不足"]]
    for route in routes:
        counts = risk_counts(route.risks)
        table_rows.append(
            [
                route.name,
                f"{route.distance_km:.1f} km",
                f"{route.duration_min} 分",
                LEVEL_LABELS[route.risk_level],
                str(route.risk_score),
                str(counts[RiskLevel.confirm_required.value]),
                str(counts[RiskLevel.data_insufficient.value]),
            ]
        )
    table = Table(table_rows, colWidths=[55 * mm, 22 * mm, 22 * mm, 28 * mm, 18 * mm, 22 * mm, 25 * mm])
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), "HeiseiMin-W3"),
                ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#eef0f2")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#c7c2b8")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 8))
    story.append(Paragraph("4. 主な注意箇所", heading_style))

    for route in routes:
        story.append(Paragraph(f"<b>{route.name}</b>", body_style))
        if not route.risks:
            story.append(Paragraph("評価済みの注意箇所はありません。ただし正式確認は必要です。", body_style))
            continue
        for risk in route.risks:
            location = ""
            if risk.feature:
                location = f" / 位置: {risk.feature.lat:.6f}, {risk.feature.lng:.6f}"
            story.append(
                Paragraph(
                    f"[{LEVEL_LABELS[risk.level]}] {risk.title}: {risk.message} "
                    f"確認先: {risk.confirmation_target}. 根拠: {risk.evidence}{location}",
                    body_style,
                )
            )

    story.append(Spacer(1, 8))
    story.append(Paragraph("5. 注意文", heading_style))
    story.append(Paragraph(DISCLAIMER, body_style))
    doc.build(story)
    return buffer.getvalue()


def render_xlsx(project: Project, routes: list[RouteCandidate]) -> bytes:
    """Render the initial-review memo as an Excel workbook (openpyxl).

    String cells are passed through ``_csv_safe`` so values that look like
    spreadsheet formulas cannot execute when the file is opened in Excel.
    """

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    header_fill = PatternFill("solid", fgColor="EEF0F2")
    header_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    overview = workbook.active
    overview.title = "概要・搬入条件"
    overview.append(["搬入ルート初期検討メモ（Excel 帳票）"])
    overview["A1"].font = Font(bold=True, size=13)
    overview.append(["案件ID", project.id])
    overview.append(["工事件名", _csv_safe(project.project_name)])
    overview.append(["現場名", _csv_safe(project.site_name)])
    overview.append(["担当者", _csv_safe(project.planner)])
    overview.append(["発注者区分", _csv_safe(project.owner_type or "未入力")])
    overview.append(
        [
            "出発地",
            f"{project.start.name} ({project.start.lat:.6f}, {project.start.lng:.6f})",
        ]
    )
    overview.append(
        [
            "到着地",
            f"{project.destination.name} ({project.destination.lat:.6f}, {project.destination.lng:.6f})",
        ]
    )
    overview.append(["車両種別", _csv_safe(project.vehicle.vehicle_type)])
    overview.append(
        [
            "全長/全幅/全高",
            (
                f"{_value(project.vehicle.length_m, 'm')} / "
                f"{_value(project.vehicle.width_m, 'm')} / {_value(project.vehicle.height_m, 'm')}"
            ),
        ]
    )
    overview.append(
        [
            "総重量/軸重",
            (
                f"{_value(project.vehicle.gross_weight_t, 't')} / "
                f"{_value(project.vehicle.axle_weight_t, 't')}"
            ),
        ]
    )
    overview.append(["積載物", _csv_safe(project.vehicle.cargo_type or "未入力")])
    overview.append(
        ["特車該当可能性", "あり" if project.vehicle.special_vehicle_flag else "未指定"]
    )
    overview.append(["搬入日", str(project.delivery.delivery_date or "未指定")])
    overview.append(["時間帯", _csv_safe(project.delivery.time_window)])
    overview.append(["回避条件", _csv_safe(", ".join(project.avoid_conditions or []) or "なし")])
    overview.append(["備考", _csv_safe(project.notes or "")])
    for row in overview.iter_rows(min_row=1, max_row=overview.max_row):
        row[0].font = header_font
    overview.column_dimensions["A"].width = 20
    overview.column_dimensions["B"].width = 90

    comparison = workbook.create_sheet("ルート候補比較")
    comparison.append(
        ["候補", "距離(km)", "時間(分)", "評価", "スコア", "注意", "要確認", "データ不足", "コメント"]
    )
    for route in routes:
        counts = risk_counts(route.risks)
        comparison.append(
            [
                _csv_safe(route.name),
                route.distance_km,
                route.duration_min,
                LEVEL_LABELS[route.risk_level],
                route.risk_score,
                counts[RiskLevel.caution.value],
                counts[RiskLevel.confirm_required.value],
                counts[RiskLevel.data_insufficient.value],
                _csv_safe(route.summary),
            ]
        )
    for cell in comparison[1]:
        cell.font = header_font
        cell.fill = header_fill
    comparison.freeze_panes = "A2"
    for column, width in zip("ABCDEFGHI", (32, 10, 10, 12, 9, 8, 9, 11, 70), strict=True):
        comparison.column_dimensions[column].width = width

    risks_sheet = workbook.create_sheet("注意箇所")
    risks_sheet.append(
        ["候補", "レベル", "種別", "名称", "内容", "確認先", "根拠", "緯度", "経度"]
    )
    for route in routes:
        for risk in route.risks:
            risks_sheet.append(
                [
                    _csv_safe(route.name),
                    LEVEL_LABELS[risk.level],
                    risk.feature.feature_type if risk.feature else "車両条件",
                    _csv_safe(risk.title),
                    _csv_safe(risk.message),
                    _csv_safe(risk.confirmation_target),
                    _csv_safe(risk.evidence),
                    round(risk.feature.lat, 6) if risk.feature else "",
                    round(risk.feature.lng, 6) if risk.feature else "",
                ]
            )
    for cell in risks_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    risks_sheet.freeze_panes = "A2"
    for column, width in zip("ABCDEFGHI", (28, 12, 12, 24, 58, 28, 42, 12, 12), strict=True):
        risks_sheet.column_dimensions[column].width = width
    for row in risks_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    notice = workbook.create_sheet("免責・注意文")
    notice.append(["本番利用禁止（PoC・サンプル）"])
    notice.append([SAMPLE_DATA_NOTICE])
    notice.append([])
    notice.append([DISCLAIMER])
    for row in notice.iter_rows():
        for cell in row:
            cell.alignment = wrap
    notice.column_dimensions["A"].width = 110

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _value(value: float | None, unit: str) -> str:
    return f"{value:g} {unit}" if value is not None else "未入力"
