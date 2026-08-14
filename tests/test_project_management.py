"""Project CRUD (edit / logical delete), list search & pagination, Excel export,
audit-log filters, and the facilities dictionary — the 2026-08-14 MVP sprint."""

from __future__ import annotations

from io import BytesIO

import pytest
from httpx import AsyncClient


def _payload(name: str = "管理機能検証 架空案件") -> dict:
    return {
        "project_name": name,
        "site_name": "管理機能検証ヤード",
        "planner": "crud-qa",
        "start": {"name": "架空出発地", "lat": 35.681236, "lng": 139.767125},
        "destination": {"name": "架空到着地", "lat": 35.658581, "lng": 139.745433},
        "vehicle": {"vehicle_type": "trailer", "length_m": 12.0, "height_m": 3.9, "gross_weight_t": 40},
        "delivery": {"delivery_date": "2026-09-15", "time_window": "daytime"},
        "avoid_conditions": ["schools"],
        "notes": "実在しない架空の検証用案件。",
    }


@pytest.mark.asyncio
async def test_list_search_filter_and_pagination(client: AsyncClient) -> None:
    for index in range(4):
        created = await client.post("/api/projects", json=_payload(f"検索対象 架空案件 {index}"))
        assert created.status_code == 201
    created = await client.post("/api/projects", json=_payload("絞込専用 架空案件"))
    assert created.status_code == 201

    listed = await client.get("/api/projects?limit=3&offset=0")
    assert listed.status_code == 200
    body = listed.json()
    assert body["total"] >= 5
    assert body["limit"] == 3
    assert body["offset"] == 0
    assert len(body["items"]) == 3
    for project in body["items"]:
        assert set(project["risk_summary"]) == {"candidates", "confirm_required", "data_insufficient"}

    page_two = await client.get("/api/projects?limit=3&offset=3")
    assert page_two.status_code == 200
    assert len(page_two.json()["items"]) > 0
    assert {p["id"] for p in page_two.json()["items"]}.isdisjoint(
        {p["id"] for p in body["items"]}
    )

    searched = await client.get("/api/projects", params={"q": "絞込専用"})
    assert searched.status_code == 200
    names = [p["project_name"] for p in searched.json()["items"]]
    assert names == ["絞込専用 架空案件"]

    by_status = await client.get("/api/projects", params={"status": "draft"})
    assert by_status.status_code == 200
    assert all(p["status"] == "draft" for p in by_status.json()["items"])


@pytest.mark.asyncio
async def test_edit_project_updates_fields_and_is_audited(
    client: AsyncClient, monkeypatch
) -> None:
    created = await client.post("/api/projects", json=_payload())
    project_id = created.json()["id"]

    updated = await client.patch(
        f"/api/projects/{project_id}",
        json={
            "project_name": "編集済み 架空案件",
            "site_name": "第二検証ヤード",
            "planner": "edit-qa",
            "start": {"name": "新出発地", "lat": 35.690000, "lng": 139.760000},
            "vehicle": {"vehicle_type": "heavy_truck", "height_m": 3.6, "gross_weight_t": 28},
            "delivery": {"time_window": "night", "night_delivery_allowed": True},
            "avoid_conditions": [],
        },
    )
    assert updated.status_code == 200, updated.text
    body = updated.json()
    assert body["project_name"] == "編集済み 架空案件"
    assert body["site_name"] == "第二検証ヤード"
    assert body["start"]["name"] == "新出発地"
    assert body["vehicle"]["gross_weight_t"] == 28
    assert body["delivery"]["night_delivery_allowed"] is True
    assert body["avoid_conditions"] == []

    fetched = await client.get(f"/api/projects/{project_id}")
    assert fetched.status_code == 200
    assert fetched.json()["project_name"] == "編集済み 架空案件"

    monkeypatch.setenv("APP_API_KEY", "admin-secret-123456")
    monkeypatch.setenv("APP_API_KEY_USER_ROLE", "admin")
    admin_headers = {"Authorization": "Bearer admin-secret-123456"}
    logs = await client.get(
        "/api/admin/audit-logs", params={"action": "project_updated"}, headers=admin_headers
    )
    assert logs.status_code == 200
    assert any(log["project_id"] == project_id for log in logs.json()["items"])


@pytest.mark.asyncio
async def test_edit_rejects_missing_and_reviewed_projects(client: AsyncClient, monkeypatch) -> None:
    created = await client.post("/api/projects", json=_payload())
    project_id = created.json()["id"]
    assert (await client.patch("/api/projects/does-not-exist", json={})).status_code == 404

    generated = await client.post(f"/api/projects/{project_id}/routes/generate", json={})
    route_id = generated.json()["routes"][0]["id"]
    assert (await client.post(f"/api/routes/{route_id}/evaluate", json={})).status_code == 200
    assert (await client.post(f"/api/projects/{project_id}/submit", json={})).status_code == 200

    monkeypatch.setenv("APP_API_KEY", "admin-secret-123456")
    monkeypatch.setenv("APP_API_KEY_USER_ROLE", "admin")
    headers = {"Authorization": "Bearer admin-secret-123456"}
    assert (
        await client.post(f"/api/projects/{project_id}/approve", json={}, headers=headers)
    ).status_code == 200

    rejected = await client.patch(
        f"/api/projects/{project_id}", json={"project_name": "承認後変更"}, headers=headers
    )
    assert rejected.status_code == 409


@pytest.mark.asyncio
async def test_delete_is_logical_archive_and_keeps_data(client: AsyncClient) -> None:
    created = await client.post("/api/projects", json=_payload("論理削除 架空案件"))
    project_id = created.json()["id"]
    generated = await client.post(f"/api/projects/{project_id}/routes/generate", json={})
    assert generated.status_code == 200

    deleted = await client.delete(f"/api/projects/{project_id}")
    assert deleted.status_code == 200
    assert deleted.json()["status"] == "archived"

    # The record and its routes survive as history; only the status changed.
    fetched = await client.get(f"/api/projects/{project_id}")
    assert fetched.status_code == 200
    assert fetched.json()["status"] == "archived"
    routes = await client.get(f"/api/projects/{project_id}/routes")
    assert routes.status_code == 200
    assert len(routes.json()) >= 3

    listed = await client.get("/api/projects", params={"status": "archived"})
    assert any(p["id"] == project_id for p in listed.json()["items"])


@pytest.mark.asyncio
async def test_xlsx_report_is_a_valid_workbook_with_notice(client: AsyncClient) -> None:
    from openpyxl import load_workbook

    created = await client.post("/api/projects", json=_payload("Excel帳票 架空案件"))
    project_id = created.json()["id"]
    generated = await client.post(f"/api/projects/{project_id}/routes/generate", json={})
    for route in generated.json()["routes"]:
        assert (
            await client.post(f"/api/routes/{route['id']}/evaluate", json={})
        ).status_code == 200

    report = await client.get(f"/api/projects/{project_id}/report?format=xlsx")
    assert report.status_code == 200, report.text
    assert report.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert report.content[:2] == b"PK"

    workbook = load_workbook(BytesIO(report.content))
    assert workbook.sheetnames == ["概要・搬入条件", "ルート候補比較", "注意箇所", "免責・注意文"]
    overview = workbook["概要・搬入条件"]
    assert overview["B3"].value == "Excel帳票 架空案件"
    comparison = workbook["ルート候補比較"]
    assert comparison.max_row >= 5
    notice = workbook["免責・注意文"]
    assert "本番利用禁止" in str(notice["A1"].value)


@pytest.mark.asyncio
async def test_xlsx_neutralizes_formula_like_cells(client: AsyncClient) -> None:
    from openpyxl import load_workbook

    payload = _payload("=HYPERLINK(\"https://example.invalid\")")
    created = await client.post("/api/projects", json=payload)
    project_id = created.json()["id"]
    await client.post(f"/api/projects/{project_id}/routes/generate", json={})
    report = await client.get(f"/api/projects/{project_id}/report?format=xlsx")
    assert report.status_code == 200
    workbook = load_workbook(BytesIO(report.content))
    cell = workbook["概要・搬入条件"]["B3"].value
    assert isinstance(cell, str)
    assert cell.startswith("'=")


@pytest.mark.asyncio
async def test_audit_logs_filters_and_export(client: AsyncClient, monkeypatch) -> None:
    monkeypatch.setenv("APP_API_KEY", "admin-secret-123456")
    monkeypatch.setenv("APP_API_KEY_USER_ID", "admin-user")
    monkeypatch.setenv("APP_API_KEY_USER_ROLE", "admin")
    headers = {"Authorization": "Bearer admin-secret-123456"}

    created = await client.post("/api/projects", json=_payload("監査フィルタ 架空案件"), headers=headers)
    project_id = created.json()["id"]
    await client.patch(
        f"/api/projects/{project_id}", json={"notes": "フィルタ確認"}, headers=headers
    )
    await client.delete(f"/api/projects/{project_id}", headers=headers)

    by_action = await client.get(
        "/api/admin/audit-logs", params={"action": "project_archived"}, headers=headers
    )
    assert by_action.status_code == 200
    assert all(log["action"] == "project_archived" for log in by_action.json()["items"])

    by_user = await client.get(
        "/api/admin/audit-logs", params={"user_id": "admin-user"}, headers=headers
    )
    assert by_user.status_code == 200
    assert by_user.json()["total"] >= 3

    by_q = await client.get(
        "/api/admin/audit-logs", params={"q": "project_updated"}, headers=headers
    )
    assert by_q.status_code == 200
    assert any(log["action"] == "project_updated" for log in by_q.json()["items"])

    exported = await client.get("/api/admin/audit-logs/export", headers=headers)
    assert exported.status_code == 200
    assert exported.headers["content-type"].startswith("text/csv")
    assert "project_archived" in exported.text


@pytest.mark.asyncio
async def test_facilities_dictionary_endpoint(client: AsyncClient) -> None:
    response = await client.get("/api/facilities")
    assert response.status_code == 200
    assert isinstance(response.json(), list)
    for point in response.json():
        assert {"id", "type", "name", "rank"} <= set(point)
